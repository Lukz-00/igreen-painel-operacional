import tempfile
import unittest
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook

from server.excel_io import preview_table, read_table, write_workbook
from server.reconciliation import reconcile_faturamento


class ReconciliationTests(unittest.TestCase):
    def test_normaliza_uc_mes_e_status(self):
        pag = pl.DataFrame({
            "Instalacao": ["000123", "456"],
            "Mês de referência": ["JAN/2026", "01/02/2026"],
            "Status": ["PAGO", "VENCIDO"],
        })
        rec = pl.DataFrame({
            "Instalacao": ["123", "000456"],
            "Data Referencia": ["2026-01", "02/2026"],
            "Status": ["PAID", "OPEN"],
        })
        result = reconcile_faturamento(pag, rec)
        self.assertEqual(result.metrics["emAmbos"], 2)
        self.assertEqual(result.metrics["coincidentes"], 1)
        self.assertEqual(result.metrics["divergentes"], 1)

    def test_cascata_numero_cliente_e_cpf(self):
        pag = pl.DataFrame({
            "Instalacao": ["10", "20"], "Mês": ["MAR/2026", "MAR/2026"],
            "CPF": ["111", "222"], "Status": ["PAGO", "PAGO"],
        })
        rec = pl.DataFrame({
            "Instalacao": ["999", "888"], "Numero Cliente": ["10", "777"],
            "Data Referencia": ["2026-03", "2026-03"], "CPF": ["000", "222"],
            "Status": ["PAID", "PAID"],
        })
        result = reconcile_faturamento(pag, rec)
        stages = set(result.sheets["COINCIDENTES"]["Etapa"].to_list())
        self.assertEqual(stages, {"2", "3"})

    def test_cpf_repetido_na_etapa_tres_nao_duplica_projecao(self):
        pag = pl.DataFrame({
            "Instalacao": ["10", "20"], "Mês": ["05/2026", "05/2026"],
            "CPF": ["123", "123"], "Nome": ["Ana", "Bruno"], "Status": ["PAGO", "PAGO"],
        })
        rec = pl.DataFrame({
            "Instalacao": ["90", "80"], "Data Referencia": ["2026-05", "2026-05"],
            "CPF": ["123", "123"], "Cliente": ["Ana", "Bruno"], "Status": ["PAID", "PAID"],
        })
        result = reconcile_faturamento(pag, rec)
        self.assertEqual(result.metrics["emAmbos"], 2)
        self.assertEqual(result.metrics["coincidentes"], 2)

    def test_duplicidades_incluem_todas_as_linhas_repetidas(self):
        pag = pl.DataFrame({"Instalacao":["1","1"], "Mês":["01/2026","01/2026"], "Status":["PAGO","PAGO"]})
        rec = pl.DataFrame({"Instalacao":["1"], "Data Referencia":["2026-01"], "Status":["PAID"]})
        result = reconcile_faturamento(pag, rec)
        self.assertEqual(result.metrics["duplicidadesPag"], 2)

    def test_clientes_gv_funciona_como_ponte_de_instalacao(self):
        pag = pl.DataFrame({"Instalacao":["900"], "Mês":["04/2026"], "Status":["PAGO"]})
        rec = pl.DataFrame({"Instalacao":["100"], "Numero Cliente":["77"], "Data Referencia":["2026-04"], "Status":["PAID"]})
        cli = pl.DataFrame({"Instalacao":["100"], "Nova Instalacao":["900"], "Numero Cliente":["77"], "Nome":["Cliente A"]})
        result = reconcile_faturamento(pag, rec, cli)
        self.assertEqual(result.metrics["emAmbos"], 1)
        self.assertEqual(result.sheets["COINCIDENTES"]["Etapa"].to_list(), ["CLI"])

    def test_openpyxl_write_only_gera_abas_validas(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "saida.xlsx"
            write_workbook(path, [("Total", 2)], {"DADOS": pl.DataFrame({"UC":["1","2"]})})
            wb = load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(wb.sheetnames, ["RESUMO", "DADOS"])
            self.assertEqual(wb["DADOS"]["A2"].value, "1")
            wb.close()

    def test_leitura_xlsx_real_por_polars(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "entrada.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Instalacao", "Mês", "Status"])
            ws.append(["00123", "JAN/2026", "PAGO"])
            wb.save(path)
            preview = preview_table(path)
            frame = read_table(path)
            self.assertEqual(preview["row_count"], 1)
            self.assertEqual(frame.shape, (1, 3))
            self.assertEqual(frame["Instalacao"][0], "00123")


if __name__ == "__main__":
    unittest.main()
