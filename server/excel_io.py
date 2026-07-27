from __future__ import annotations

import csv
import re
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _unique_headers(values: Iterable[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, 1):
        base = str(value).strip() if value not in (None, "") else f"COLUNA_{index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def preview_table(path: Path, limit: int = 3, sheet_name: str | None = None) -> dict[str, Any]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
            sample = fh.read(131072)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(fh, dialect)
            headers = _unique_headers(next(reader, []))
            rows = [dict(zip(headers, row)) for _, row in zip(range(limit), reader)]
        # Contagem em streaming, sem materializar a planilha.
        with path.open("rb") as fh:
            row_count = max(sum(chunk.count(b"\n") for chunk in iter(lambda: fh.read(8 << 20), b"")) - 1, 0)
        return {"headers": headers, "rows": rows, "row_count": row_count, "sheets": [], "sheet_name": ""}

    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        selected = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[selected]
        iterator = ws.iter_rows(values_only=True)
        headers = _unique_headers(next(iterator, []))
        rows = [dict(zip(headers, values)) for _, values in zip(range(limit), iterator)]
        if ws.max_row is None:
            ws.calculate_dimension(force=True)
        return {"headers": headers, "rows": rows, "row_count": max((ws.max_row or 1) - 1, 0), "sheets": wb.sheetnames, "sheet_name": selected}
    finally:
        wb.close()


def read_table(path: Path, sheet_name: str | None = None) -> pl.DataFrame:
    """Lê CSV/XLSX com Polars; todos os campos ficam textuais e previsíveis."""
    if path.suffix.lower() == ".csv":
        return pl.read_csv(
            path,
            infer_schema=False,
            ignore_errors=True,
            encoding="utf8-lossy",
            separator=_detect_separator(path),
            null_values=[],
            truncate_ragged_lines=True,
        ).fill_null("")
    frame = pl.read_excel(path, engine="calamine", infer_schema_length=0, sheet_name=sheet_name) if sheet_name else pl.read_excel(path, engine="calamine", infer_schema_length=0)
    return frame.select(pl.all().cast(pl.String, strict=False).fill_null(""))


def _detect_separator(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(65536)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ";"


_INVALID_SHEET = re.compile(r"[\\/*?:\[\]]")


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = _INVALID_SHEET.sub("_", name).strip(" '")[:31] or "DADOS"
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        tail = f"_{suffix}"
        candidate = base[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _set_data_widths(sheet: Any, frame: pl.DataFrame, sample_size: int = 200) -> None:
    sample = frame.head(sample_size)
    wide_fields = ("arquivo", "barras", "motivo", "origem", "meses")
    for index, column in enumerate(frame.columns, 1):
        values = sample.get_column(column).to_list() if sample.height else []
        longest = max([len(str(column)), *(len(str(value)) for value in values if value is not None)])
        cap = 48 if any(token in str(column).casefold() for token in wide_fields) else 32
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), cap)


def write_workbook(path: Path, summary: list[tuple[str, Any]], sheets: Mapping[str, pl.DataFrame]) -> None:
    """Gera XLSX em write-only: memória constante mesmo com centenas de milhares de linhas."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)
    used: set[str] = set()
    ws = wb.create_sheet(safe_sheet_name("RESUMO", used))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 72
    for index, (label, value) in enumerate(summary):
        label_cell = WriteOnlyCell(ws, value=label)
        value_cell = WriteOnlyCell(ws, value=_excel_value(value))
        if index == 0:
            for cell in (label_cell, value_cell):
                cell.fill = header_fill
                cell.font = header_font
        else:
            label_cell.fill = label_fill
            label_cell.font = Font(bold=True, color="1F2937")
        ws.append([label_cell, value_cell])

    for title, frame in sheets.items():
        if frame.height == 0:
            continue
        out = wb.create_sheet(safe_sheet_name(title, used))
        _set_data_widths(out, frame)
        out.row_dimensions[1].height = 34
        header = []
        for value in frame.columns:
            cell = WriteOnlyCell(out, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header.append(cell)
        out.append(header)
        out.freeze_panes = "A2"
        out.auto_filter.ref = f"A1:{get_column_letter(len(frame.columns))}1"
        for row in frame.iter_rows():
            out.append([_excel_value(value) for value in row])
    wb.save(path)


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    return str(value)
